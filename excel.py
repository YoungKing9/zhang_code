from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from xlwt import *
import xlwt
import numpy as np


def make_excel():
    path = '/data/WorkMind/data/code/project_lab/github/data/小广场高光_img_log_20/log.txt'
    save_excel = 'ex.xlsx'
    tx = open(path, 'r').readlines()
    wb = Workbook()
    # ws = wb.create_sheet(title='dts', index=0)
    row = 10
    col = 2
    # 1614044732143463_L.png: 173, 122, 74

    for i in tx:
        da = i.split(':')[1]
        print(da)
        # _ = ws.cell(colum=col, row=row, value="{0}".format(get_column_letter(col))).value = da.split(',')[0]
        # col += 1
    #
    # wb.save(save_excel)




def make_xls():
    path = '/data/WorkMind/data/code/project_lab/github/data/小广场高光_img_log_20/log.txt'
    tx = open(path, 'r').readlines()

    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('记录数据',cell_overwrite_ok=True)
    # 335, 324, 204
    for i, data in enumerate(tx): # 一共10条数据，也就是10行
        da = data.split(':')[1]
        for j in range(3):
            sheet.write(i, j, int(da.split(',')[j]))
    take_var(sheet, tx)
    savepath = '/data/WorkMind/data/code/project_lab/github/data/take_data.xlsx'
    book.save(savepath)




def take_var(sheet,tx):
    data_lsit1 = []
    data_lsit2 = []
    for i, data in enumerate(tx): # 一共10条数据，也就是10行
        da = data.split(':')[1]
        data_lsit1.append(int(da.split(',')[0]))
        data_lsit2.append(int(da.split(',')[2]))

    max1 = int(np.max(data_lsit1))
    max2 = int(np.max(data_lsit2))

    min1 = int(np.min(data_lsit1))
    min2 = int(np.min(data_lsit2))

    std1 = round(float(np.std(data_lsit1)), 2)
    std2 = round(float(np.std(data_lsit2)), 2)

    mean1 = round(float(np.mean(data_lsit1)), 2)
    mean2 = round(float(np.mean(data_lsit2)), 2)

    scale1 = round(float(std1)/float(mean1), 2)
    scale2 = round(float(std2)/float(mean2), 2)

    data_li1 = [max1, min1, std1, mean1, scale1]
    data_li2 = [max2, min2, std2, mean2, scale2]
    print(data_li1)
    print(data_li2)

    for ii in range(3):
        if ii == 0:
            for jj in range(5):
                sheet.write(10 + jj, ii, data_li1[jj])
        if ii == 2:
            for jj in range(5):
                sheet.write(10 + jj, ii, data_li1[jj])


def make_use_excel():
    path = '/data/WorkMind/data/code/project_lab/github/data/小广场高光_img_log_20/log.txt'
    save_excel = 'ex.xlsx'
    tx = open(path, 'r').readlines()

    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('豆瓣电影Top250',cell_overwrite_ok=True)
    col = ('电影详情链接', '图片链接', '影片中文名', '影片外国名', '评分', '评价数', '概况', '相关信息')
    # for i in range(0, 8):
    #     sheet.write(0, i, col[i])
    datalist = [['www', 'www图片', '西游记', 'xiyouji', '100分', '0人', '很好', '超级棒'],
                ['www2', 'www图片2', '西游记2', 'xiyouji2', '1000分', '1人', '很棒', '一级棒']]
    # for i in range(0, 2):
    #     data = datalist[i]
    #     for j in range(0, 8):
    #         sheet.write(i + 1, j, data[j])
    # 335, 324, 204
    for i, data in enumerate(tx): # 一共10条数据，也就是10行
        da = data.split(':')[1]
        for j in range(3):
            sheet.write(i, j, da.split(',')[j])
    savepath = 'excel表格.xls'
    book.save(savepath)

if __name__ == '__main__':
    # make_excel()
    make_xls()
